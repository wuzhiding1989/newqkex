from openpyxl import load_workbook
from common import slacksend
#coding:utf-8
import pygsheets
import openpyxl
from BU.futures.api import webapi as web
#user = wb.webapi(5, 'test')
excel_path = "cccda12.xlsx"
sheet_name = "计算";tradeType='linearPerpetual';gear='depth-3';limit=1000;Price=1800;symbol='ETHUSDT'

def update_data(symbol,Price):#将用户的资产，交易对配置等填充到表格
    user = web.webapi(5, 'test')
    res=user.web_instruments(tradeType=tradeType,symbol=symbol)
    takerRate=res['data'][0]['takerRate'];makerRate=res['data'][0]['makerRate']
    markPriceGreaterRatio = res['data'][0]['markPriceGreaterRatio'];maintMarginRatio = res['data'][0]['maintMarginRatio']
    res2=user.web_tradingAccount(currency='USDT')
    marginAvailable = res2['data'][0]['marginAvailable'];marginFrozen = res2['data'][0]['marginFrozen']
    marginEquity = res2['data'][0]['marginEquity'];marginPosition=res2['data'][0]['marginPosition']
    ak = user.web_market_depth(tradeType=tradeType, gear=gear, symbol=symbol, limit=limit)
    buy1 = ak['data']['bids'][0][0];buy1=f'{buy1}'
    sell1 =ak['data']['asks'][0][0];sell1=f'{sell1}'
    res1=user.web_position(tradeType=tradeType,symbol=symbol,marginType='cross')
    # if len(res1['data']) != 0:
    #     for i, item in enumerate(res1['data']):
    #         trade_type = item["tradeType"]
    #         var_name = f"tradeType{i + 1}"
    #         exec(f"{var_name} = '{trade_type}'"))
    # else:
    #     print(symbol)
    avgEntryPrice=res1['data'][0]['avgEntryPrice'];leverage=res1['data'][0]['leverage']
    markPrice=res1['data'][0]['markPrice'];liquidationPrice=res1['data'][0]['liquidationPrice']
    posMargin = res1['data'][0]['posMargin'];marginRate=res1['data'][0]['marginRate']
    unrealisedPnl = res1['data'][0]['unrealisedPnl'];positionAmt = res1['data'][0]['positionAmt']
    side = res1['data'][0]['side'];earningRate = res1['data'][0]['earningRate']
    # 读取 Excel 表格数据
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    # 更新 data1 和 data2 属性值
    ws["A2"] = symbol;ws["B6"] = symbol;ws["A11"] = symbol;ws["B8"] = symbol
    ws["B2"] = Price
    ws["C2"] = avgEntryPrice;ws["D2"] = side;ws["E2"] = leverage;ws["H2"] = positionAmt ;ws["L2"] = markPrice
    ws["E21"] = maintMarginRatio
    ws["F2"] = 26040 #avgPrice
    ws["G2"] = takerRate
    ws["B21"] = marginEquity#汇总资产
    ws["C21"] = marginAvailable;ws["F21"] = marginPosition
    ws["D21"] = marginFrozen
    ws["M2"] = buy1
    ws["N2"] = sell1
    ws["O2"] = markPriceGreaterRatio
    ws['p2'] = makerRate

    ws['C8'] = liquidationPrice;ws['D8'] = posMargin;ws['E8'] = marginRate;ws['F8'] = None#维持保证金
    ws['H8'] = unrealisedPnl
    ws['I8'] = earningRate
    #ws['J8'] = earningRate
    # 将数据写入 Excel 表格
    wb.save(excel_path)
    ############写入完数据后需要手动保存一下文件，否则无法生效


def select_data(symbol):  # 将用户的资产，交易对配置等填充到表格
    user = web.webapi(2, 'test')
    res = user.web_instruments(tradeType=tradeType, symbol=symbol)
    takerRate = res['data'][0]['takerRate'];
    markPriceGreaterRatio = res['data'][0]['markPriceGreaterRatio'];
    res2 = user.web_tradingAccount(currency='USDT')
    marginAvailable = res2['data'][0]['marginAvailable'];
    marginFrozen = res2['data'][0]['marginFrozen']
    marginEquity = res2['data'][0]['marginEquity'];
    marginPosition = res2['data'][0]['marginPosition']
    res1 = user.web_position(tradeType=tradeType, symbol=symbol, marginType='cross')
    if len(res1['data'])==0:
        return
    if len(res1['data']) == 2:
        avgEntryPrice1 = res1['data'][1]['avgEntryPrice'];
        symbol1 = res1['data'][1]['symbol'];
        leverage1 = res1['data'][1]['leverage']
        markPrice1 = res1['data'][1]['markPrice'];
        side1 = res1['data'][1]['side'];
        positionAmt1 = res1['data'][1]['positionAmt']
        avgEntryPrice = res1['data'][0]['avgEntryPrice'];
        leverage = res1['data'][0]['leverage']
        symbol = res1['data'][0]['symbol'];
        markPrice = res1['data'][0]['markPrice'];
        positionAmt = res1['data'][0]['positionAmt']
        side = res1['data'][0]['side'];
        wb = load_workbook(excel_path)
        ws = wb['t1']
        ws["D5"] = avgEntryPrice1;ws["S5"] = side1;ws["T5"] = positionAmt1;ws["F5"] = leverage1;ws["C5"] = markPrice1;ws["A5"] = symbol1;
        ws["D4"] = avgEntryPrice;
        ws["S4"] = side;
        ws["T4"] = positionAmt;
        ws["F4"] = leverage;
        ws["C4"] = markPrice;
        ws["J2"] = marginEquity  # 汇总资产
        ws["G2"] = marginAvailable;
        ws["A4"] = symbol;
        ws["I2"] = marginPosition
        ws["H2"] = marginFrozen
        ws["H4"] = takerRate;
        ws["H5"] = takerRate
        ws["I4"] = markPriceGreaterRatio;
        ws["I5"] = markPriceGreaterRatio
        # 将数据写入 Excel 表格
        wb.save(excel_path)
    elif len(res1['data']) == 1:
        avgEntryPrice = res1['data'][0]['avgEntryPrice'];
        leverage = res1['data'][0]['leverage']
        symbol = res1['data'][0]['symbol'];
        markPrice = res1['data'][0]['markPrice'];
        positionAmt = res1['data'][0]['positionAmt']
        side = res1['data'][0]['side'];
        # 读取 Excel 表格数据
        wb = load_workbook(excel_path)
        ws = wb['t1']
        # 更新 data1 和 data2 属性值
        ws["D4"] = avgEntryPrice;
        ws["S4"] = side;
        ws["T4"] = positionAmt;
        ws["F4"] = leverage;
        ws["C4"] = markPrice;
        ws["J2"] = marginEquity  # 汇总资产
        ws["G2"] = marginAvailable;
        ws["A4"] = symbol;
        ws["I2"] = marginPosition
        ws["H2"] = marginFrozen
        ws["H4"] = takerRate;ws["H5"] = takerRate
        ws["I4"] = markPriceGreaterRatio;ws["I5"] = markPriceGreaterRatio
        ws["T5"] = 0
        # 将数据写入 Excel 表格
        wb.save(excel_path)
        ############写入完数据后需要手动保存一下文件，否则无法生效
    else:
        print(len(res1['data']))
        return

def str_add_one(s):
    # 将字符串最后一位字符转成 ASCII 码
    ascii_code = ord(s[-1])
    # 判断如果是 '9'，则进位到 'A'
    if ascii_code == 57:
        return s[:-1] + 'A'
    # 判断如果是 'Z'，则进位到 'AA'
    elif ascii_code == 90:
        return str_add_one(s[:-1]) + 'A'
    # 其他情况直接将最后一位字符+1即可
    else:
        return s[:-1] + chr(ascii_code + 1)

def str_add_n(s, n):
    # 将字符串最后一位字符转成 ASCII 码
    ascii_code = ord(s[-1])
    # 判断如果是 '9'，则进位到 'A'
    if ascii_code == 57:
        return str_add_n(s[:-1], n) + 'A'
    # 判断如果是 'Z'，则进位到 'AA'
    elif ascii_code == 90:
        return str_add_n(s[:-1], n+1) + 'A'
    # 其他情况直接将最后一位字符+n即可
    else:
        new_char = chr(ascii_code + n)
        # 如果新字符是 '4'，将其替换为 '6'
        if new_char == '4':
            new_char = '6'
        return s[:-1] + new_char

def java_read_cell_value():#后端计算相关值
    # 读取 Excel 表格数据
    book = load_workbook(excel_path,data_only=True)
    sheet = book[sheet_name]
    result = []
    ac = ['A5','B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5']
    for i in range(len(ac)):
        acc=str_add_one(ac[i])
        #读取结果的单元格值（假设结果单元格为 A1）
        result_1 = sheet[ac[i]].value
        result_2 = sheet[acc].value
        result.append('%s:%s' % (result_1, result_2))
        #print(result_1,result_2,result)
    return result

def api_read_cell_value():#打印后端计算的值
    # 读取 Excel 表格数据
    book = load_workbook(excel_path,data_only=True)
    sheet = book[sheet_name]
    result = []
    ac = ['A5','B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5']
    for i in range(len(ac)):
        acc=str_add_n(ac[i],3)
        #读取结果的单元格值（假设结果单元格为 A1）
        result_1 = sheet[ac[i]].value
        result_2 = sheet[acc].value
        result.append('%s:%s' % (result_1, result_2))
        #print(result_1,result_2,result)
    return result
def web_read_cell_value():#前段计算值
    # 读取 Excel 表格数据
    book = load_workbook(excel_path,data_only=True)
    sheet = book[sheet_name]
    result = []
    ac = ['A10','B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10']
    for i in range(len(ac)):
        acc=str_add_one(ac[i])
        #读取结果的单元格值（假设结果单元格为 A1）
        result_1 = sheet[ac[i]].value
        result_2 = sheet[acc].value
        result.append('%s:%s' % (result_1, result_2))
        #print(result_1,result_2,result)
    return result





if __name__ == '__main__':
    #print(update_data(symbol,Price))
    # slacksend.send_Slack(java_read_cell_value())
    # slacksend.send_Slack(web_read_cell_value())
    print(select_data(symbol))
    # print(api_read_cell_value())
    # print(web_read_cell_value())
